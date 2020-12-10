import openpyxl
import requests
import time


class CopyExcel:
    def __init__(self, src):
        self.wb = openpyxl.load_workbook(src)
        self.ws = self.wb.worksheets[0]
        self.dest = src

    def write_workbook(self, row_dest, column_dest, value):
        c = self.ws.cell(row=row_dest, column=column_dest)
        c.value = value

    def save_excel(self):
        self.wb.save(self.dest)


filename = 'file_path'
copy = CopyExcel(filename)
treking_column = None
for col in range(copy.ws.min_column, copy.ws.max_column):
    if copy.ws.cell(1, col).value == 'Numer trekingowy':
        treking_column = col
        break

treking_numbers = []
for row in range(2, copy.ws.max_row):
    if type(copy.ws.cell(row, treking_column).value) == int:
        if copy.ws.cell(row, 12).value != "The shipment has been successfully delivered":
            treking_numbers.append(copy.ws.cell(row, treking_column).value)

headers = {'Accept': 'application/json', 'DHL-API-Key': 'API-KEY'}
for number in treking_numbers:
    r = requests.get('https://api-eu.dhl.com/track/shipments?trackingNumber={}&language=en&limit=1'.format(number),
                     headers=headers)
    try:
        r.json()['status']
    except KeyError:
        status = r.json()['shipments'][0]['status']['description']
        date = r.json()['shipments'][0]['status']['timestamp'].split('T')[0]
        data = {'number': number, 'status': status, 'date': date}
        for row in range(2, copy.ws.max_row):
            if copy.ws.cell(row, treking_column).value == data['number']:
                copy.write_workbook(row_dest=row, column_dest=12, value=data['status'])
                copy.write_workbook(row_dest=row, column_dest=13, value=data['date'])
                copy.save_excel()
                break
    time.sleep(2)
    print('{}/{}'.format(treking_numbers.index(number) + 1, len(treking_numbers)))
